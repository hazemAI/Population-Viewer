# This program displays population of some countries and provinces with excel databases.
# Author: Hazem Khaled (FCAI student).
# Date: 8 Mar. 2022
# Version: 1.0

from openpyxl import Workbook, load_workbook
# try except block to accept only valid inputs.
while True:  
    try:
        wb = load_workbook('Population.xlsx')
        firstChoice = int(input('1-Egypt\n2-USA\n3-Canada\n4-Exit\nplease enter country number to choose it or 4 to exit: '))
        

        if firstChoice == 1:
            while True:  
                try:
                    ws = wb['Egypt'] # loads Egypt's worksheet.
                    secondChoice = int(input(('1- display all provinces population and the country\'s total population\n\
2- display the province with highest population and the one with the lowest\n3- back\n')))
                    if secondChoice == 1:
                        rows = ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=27)
                        print('Egypt : 92,000,000')
                        for a,b in rows:
                            print(a.value, b.value) # displays all provinces/states population with the entire country.

                    elif secondChoice == 2:
                        ws = wb['Egypt']
                        cell1 = ws.cell(row = 1, column = 1)
                        cell2 = ws.cell(row = 1, column = 2)
                        cell3 = ws.cell(row = 27, column = 1)
                        cell4 = ws.cell(row = 27, column = 2)
                        # displays the first and last provinces/state in database which have the highest and lowest population.
                        print(f'\n{cell1.value} : {cell2.value}\n{cell3.value} : {cell4.value}\n')

                    elif secondChoice == 3:
                        break


                    else:
                        print("\n***invalid input***\n")        
                
                except ValueError:
                    print('\n***invalid input***\n')


        elif firstChoice == 2:
            while True:  
                try:
            
                    ws = wb['USA']
                    secondChoice = int(input(('1- display all provinces population and the country\'s total population\n\
2- display the province with highest population and the one with the lowest\n3- back\n')))
                    if secondChoice == 1:
                        rows = ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=50)
                        print('USA : 329,450,000')
                        for a,b in rows:
                            print(a.value, b.value)

                    elif secondChoice == 2:
                        ws = wb['USA']
                        cell1 = ws.cell(row = 1, column = 1)
                        cell2 = ws.cell(row = 1, column = 2)
                        cell3 = ws.cell(row = 50, column = 1)
                        cell4 = ws.cell(row = 50, column = 2)
                        print(f'\n{cell1.value} : {cell2.value}\n{cell3.value} : {cell4.value}\n')

                    elif secondChoice == 3:
                        break

                    else:
                        print("\n***invalid input***\n")

                except ValueError:
                    print('\n***invalid input***\n')
            
        elif firstChoice == 3:
            while True:
                try:
                    ws = wb['Canada']
                    secondChoice = int(input(('1- display all provinces population and the country\'s total population\n\
2- display the province with highest population and the one with the lowest\n3- back\n')))
                    if secondChoice == 1:
                        rows = ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=13)
                        print('Canada : 38,000,000')
                        for a,b in rows:
                            print(a.value, b.value)

                    elif secondChoice == 2:
                        ws = wb['Canada']
                        cell1 = ws.cell(row = 1, column = 1)
                        cell2 = ws.cell(row = 1, column = 2)
                        cell3 = ws.cell(row = 13, column = 1)
                        cell4 = ws.cell(row = 13, column = 2)
                        print(f'\n{cell1.value} : {cell2.value}\n{cell3.value} : {cell4.value}\n')

                    elif secondChoice == 3:
                        break

                    else:
                        print("\n***invalid input***\n")

                except ValueError:
                    print("\n***invalid input***\n")

        elif firstChoice == 4:
            print("Thank you for using the app!")
            quit()

        else:
            print("\n***invalid input***\n")

    except ValueError:
        print('\n***invalid input***\n')
